import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def leer_excel(archivo):
    """
    Lee un archivo Excel completo y lo devuelve como un DataFrame.
    """
    try:
        df = pd.read_excel(archivo)
        return df
    except Exception as e:
        print(f"Error al leer el archivo {archivo}: {e}")
        return None

def encontrar_coincidencias(df1, df2):
    """
    Busca coincidencias entre las filas de dos DataFrames, sin importar su posición.
    Devuelve un DataFrame fusionado con las coincidencias marcadas.
    """
    # Realizar la comparación sin importar el orden de las filas
    df1['source'] = 'file1'
    df2['source'] = 'file2'
    concatenado = pd.concat([df1, df2])

    # Eliminar duplicados basados en todas las columnas excepto 'source'
    duplicados = concatenado.drop_duplicates(subset=concatenado.columns.difference(['source']), keep=False)

    # Crear un dataframe fusionado para mostrar las coincidencias
    merged = pd.merge(df1, df2, how='outer', indicator=True, on=df1.columns.difference(['source']).tolist())
    return merged, duplicados

def comparar_archivos(archivo1, archivo2, reporte_salida):
    """
    Compara dos archivos Excel y genera un reporte con la comparación fila por fila,
    permitiendo coincidencias en diferentes posiciones.
    """
    df1 = leer_excel(archivo1)
    df2 = leer_excel(archivo2)

    if df1 is None or df2 is None:
        print("No se pudieron leer ambos archivos. Revisa las rutas o el formato.")
        return

    # Comparar archivos
    merged, duplicados = encontrar_coincidencias(df1, df2)

    # Crear el archivo Excel para los resultados
    wb = Workbook()

    # Hoja 1: Comparación combinada con colores
    ws1 = wb.active
    ws1.title = "Comparación Combinada"

    # Estilos y colores
    fill_match = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Verde para coincidencias
    fill_diff = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Rojo para diferencias
    fill_unique = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Naranja para valores únicos
    fill_header = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Verde claro para encabezados
    bold_font = Font(bold=True)  # Negrita para encabezados

    # Insertar encabezados en hoja 1
    for col_num, column in enumerate(merged.columns, 1):
        cell = ws1.cell(row=1, column=col_num, value=column)
        cell.font = bold_font
        cell.fill = fill_header

    # Insertar valores en hoja 1 y aplicar colores según la comparación
    for row_num, row in merged.iterrows():
        for col_num, value in enumerate(row, 1):
            cell = ws1.cell(row=row_num + 2, column=col_num, value=value)

            if row['_merge'] == 'both':  # Coincidencias en ambas tablas
                cell.fill = fill_match
            elif row['_merge'] == 'left_only':  # Solo en archivo1
                cell.fill = fill_diff
            elif row['_merge'] == 'right_only':  # Solo en archivo2
                cell.fill = fill_diff
            else:
                cell.fill = fill_unique

    # Hoja 2: Reporte de filas revisadas
    ws2 = wb.create_sheet(title="Reporte Revisado")
    for col_num, column in enumerate(duplicados.columns, 1):
        cell = ws2.cell(row=1, column=col_num, value=column)
        cell.font = bold_font
        cell.fill = fill_header

    for row_num, row in duplicados.iterrows():
        for col_num, value in enumerate(row, 1):
            ws2.cell(row=row_num + 2, column=col_num, value=value)

    # Guardar el archivo
    wb.save(reporte_salida)
    print(f"Reporte generado y guardado como '{reporte_salida}'.")

if __name__ == "__main__":
    archivo1 = 'file1.xlsx'  # Cambia esto a la ruta de tu archivo 1
    archivo2 = 'file2.xlsx'  # Cambia esto a la ruta de tu archivo 2
    reporte_salida = 'comparacion_avanzada.xlsx'  # El nombre del archivo de salida

    comparar_archivos(archivo1, archivo2, reporte_salida)
